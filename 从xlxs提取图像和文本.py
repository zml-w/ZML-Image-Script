import os
from openpyxl import load_workbook

# 定义一个函数，用于从Excel中提取文本和图片
def extract_excel_by_column(
    file_path,       # Excel文件的路径
    sheet_name,      # 要操作的工作表名称
    text_cols,       # 包含文本的列的列表，例如 ['A', 'D']
    image_cols,      # 包含图片的列的列表，例如 ['B', 'E']
    start_row,       # 开始处理的行号
    output_dir='output' # 输出文件夹的名称，默认为 'output'
):
    # 使用 load_workbook 加载 Excel 文件。
    # data_only=True 表示只加载单元格的显示值，而不是公式。
    wb = load_workbook(file_path, data_only=True)

    # 检查用户输入的工作表名称是否存在
    if sheet_name not in wb.sheetnames:
        print("可用的工作表：", wb.sheetnames) # 提示用户所有可用的工作表名称
        raise ValueError(f'工作表 {sheet_name} 不存在') # 抛出错误，终止程序

    # 获取指定的工作表
    ws = wb[sheet_name]
    
    # 创建输出目录。exist_ok=True 意味着如果目录已经存在，则不会报错。
    os.makedirs(output_dir, exist_ok=True)

    # ===== 1. 每一列单独保存文本 =====
    # 遍历所有需要提取文本的列
    for col in text_cols:
        # 构造文本文件的完整路径，例如 'output/D.txt'
        txt_path = os.path.join(output_dir, f"{col}.txt")
        
        # 以写入模式 ('w') 打开文件，并指定编码为 'utf-8-sig'，
        # 'utf-8-sig' 可以保证在Windows记事本中打开txt文件时不会出现乱码。
        with open(txt_path, 'w', encoding='utf-8-sig') as f:
            # 遍历从 start_row 到 Excel最大行 的所有行
            for row in range(start_row, ws.max_row + 1):
                # 获取当前单元格的值，例如 ws['D7'] 的值
                val = ws[f"{col}{row}"].value
                # 如果单元格有值（不是 None），则写入文件
                if val is not None:
                    # 将值转换成字符串并写入文件，并在末尾添加一个换行符
                    f.write(str(val) + "\n")
        print(f"文本列 {col} 已保存到: {txt_path}") # 打印提示信息

    # ===== 2. 每一列单独保存图片 =====
    # 注意：openpyxl 将所有图片存储在 ws._images 列表中，而不是按单元格存储。
    
    # 提前为每个图片列创建其子目录，避免在循环中重复创建
    for col in image_cols:
        img_dir = os.path.join(output_dir, col)
        os.makedirs(img_dir, exist_ok=True)

    # 遍历工作表中所有的图片对象
    for image in ws._images:
        # 获取图片在Excel中的锚点信息（即图片所关联的单元格位置）
        anchor = image.anchor._from
        
        # openpyxl 的锚点是基于0的索引，所以需要加1来转换为Excel的基于1的行和列号
        # 获取图片所在的列字母，例如 'E'
        col_letter = ws.cell(row=anchor.row + 1, column=anchor.col + 1).column_letter
        # 获取图片所在的行号，例如 7
        row_num = anchor.row + 1

        # 检查图片是否在用户指定的图片列中，并且行号大于等于起始行
        if col_letter in image_cols and row_num >= start_row:
            # 构造图片保存的文件夹路径，例如 'output/E'
            img_folder = os.path.join(output_dir, col_letter)
            # 确保图片文件夹存在 (这里因为前面已经提前创建，所以这一步也可以省略，但保留更安全)
            os.makedirs(img_folder, exist_ok=True)
            
            # 构造图片的完整保存路径，例如 'output/E/7.png'
            img_path = os.path.join(img_folder, f"{row_num}.png")
            
            # 以二进制写入模式 ('wb') 打开文件
            with open(img_path, 'wb') as img_file:
                # 将图片数据写入文件
                img_file.write(image._data())
            print(f"图片 {col_letter}{row_num} 已保存到: {img_path}") # 打印提示信息

    print("\n提取完成！") # 所有操作完成后打印提示

# 判断当前文件是否是直接运行的（而不是被import的）
if __name__ == '__main__':
    # ===== 用户交互部分 =====

    # 提示用户拖拽 Excel 文件，并去除可能存在的引号和空格
    excel_path = input("请将 .xlsx 文件拖入此窗口后回车：\n").strip().strip('"')

    # 用于获取工作表名称，只读模式，不加载数据，提高效率
    tmp_wb = load_workbook(excel_path, read_only=True)
    print("\n可用工作表：", tmp_wb.sheetnames)
    sheet_name = input("请输入要提取的工作表名称：").strip()

    # 提示用户输入文本列，并进行处理：去除空格，按逗号分割成列表
    text_cols = input("请输入要提取的文本列(逗号分隔，如 D,G,J)：").replace(' ', '').split(',')
    # 提示用户输入图片列，并进行处理：去除空格，按逗号分割成列表
    image_cols = input("请输入要提取的图片列(逗号分隔，如 E,H,K)：").replace(' ', '').split(',')

    # 提示用户输入起始行号，并转换为整数
    start_row = int(input("请输入开始提取的行号(例如 7)：").strip())

    # 提示用户输入输出目录，如果用户直接回车，则使用默认值 'output'
    output_dir = input("输出目录（默认 output，可直接回车）：").strip() or 'output'

    # 调用主函数执行提取操作
    extract_excel_by_column(excel_path, sheet_name, text_cols, image_cols, start_row, output_dir)

    # 在所有功能执行完毕后，暂停程序，等待用户按下回车键，这样命令行窗口就不会立即关闭。
    input("\n所有操作已完成，按回车键退出...")
